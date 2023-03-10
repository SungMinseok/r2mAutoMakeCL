import pandas as pd
import time
import os
import gc
import re
import numpy as np
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Color
from tqdm import tqdm
from openpyxl.styles.colors import BLACK
import datetime
from openpyxl.styles import Font, Color
#from openpyxl.styles.colors import RED
from openpyxl.utils import get_column_letter
#cashShopIdIndexList = cashShopIdList.index
clType = ""
cashShopDir = ""
tempDir = "./temp"
if not os.path.isdir(tempDir) :
    os.mkdir(tempDir)
tempCsvName = f"./temp/tempCsv.csv"

xlFileName = ""
tcStartDate = ""

def dateCheck(start_date, end_date, today = datetime.date.today()):
    
    if start_date.date() == today :
        return "이벤트 시작"
    elif start_date.date() < today < end_date.date():
        return "이벤트 유지"
    elif today == end_date.date():
        return "이벤트 종료"
    # elif start_date.date() >= today - datetime.timedelta(days=1)  :
    #     return "판매 전"
    elif start_date.date() >= today  :
        return "이벤트 시작 전"
    else:
        return "이벤트 제외"

class Event():

    
    def __init__(self) :

        self.id = ""
        self.eType = ""
        self.name = ""
        self.questDesc = ""

        self.itemList0 = []
        self.itemList1 = []
        self.itemList2 =[]

        self.craftPrice = ""
        self.craftIngred = ""
        self.limit = ""
        self.server = ""
        self.startDate = ""
        self.endDate = ""

        #별도 저장값
        self.startCheck = ""

def extract_data_event(fileName):

#CSV 읽기
    #target = pd.read_csv(fileName)

#XLSX 읽기
    tempTarget = pd.read_excel(fileName,engine='openpyxl', na_values = "")
    tempTarget.to_csv(tempCsvName, encoding='cp949')
    target = pd.read_csv(tempCsvName, encoding='cp949')

    target = target.replace('-',np.nan)
    idList = target["EventID"].dropna(axis=0)
    idIndexList = idList.index

    totalCount = len(idIndexList)

    eventList = [Event]
    eventList.clear()

    print("데이터 추출 중...")
    for j in tqdm(range(0,totalCount)):

        if (j+1) >= len(idIndexList) :
            tempDf = target[idIndexList[j]:]
        else :
            tempDf = target[idIndexList[j]:idIndexList[j+1]]
        tempDf = tempDf.reset_index()

        a = Event()
        a.id = int(tempDf.loc[0,"CashShopID"])
        a.pkgName = tempDf.loc[0,"PkgName"] #+ "[귀속]"
        a.category = tempDf.loc[0,"Category"]
        a.price = str(tempDf.loc[0,"Price"])
        try:
            a.bonus = int(tempDf.loc[0,"Bonus"])
        except:
            a.bonus = 0
        a.limit = tempDf.loc[0,"Limit"]

        for k in range(len(tempDf)):
            #print(len(tempDf))
            if not pd.isnull(tempDf.iloc[k]['Name0']):
                itemName = tempDf.iloc[k]['Name0']
                itemCount = tempDf.iloc[k]['Count0']
                try : 
                    a.itemList0.append(f"{itemName}[귀속] {int(itemCount)}개")
                except:
                    a.itemList0.append(f"{itemName}[귀속] {(itemCount)}개")
                #print(a.itemList0)

        for k in range(len(tempDf)):
            if not pd.isnull(tempDf.iloc[k]['Name1']):
                itemName = tempDf.iloc[k]['Name1']
                itemCount = tempDf.iloc[k]['Count1']
                try: 
                    a.itemList1.append(f"{itemName}[귀속] {int(itemCount)}개")
                except:
                    a.itemList1.append(f"{itemName}[귀속] {(itemCount)}개")

        a.server = tempDf.loc[0,"Server"]
       
        a.startDate = pd.to_datetime(tempDf.loc[0,"StartDate"])

        if "상시" in str(tempDf.loc[0,"EndDate"]) :
            a.endDate = datetime.datetime.strptime("2099-12-31 00:00:00",'%Y-%m-%d %H:%M:%S')
        else: 
            a.endDate = pd.to_datetime(tempDf.loc[0,"EndDate"])
          
        if fileType == "0" :#TC
            startDate = datetime.datetime.strptime(tcStartDate, '%Y-%m-%d')
            a.salesCheck = dateCheck(a.startDate,a.endDate,startDate.date())
        elif fileType == "1" :#점검
            a.salesCheck = dateCheck(a.startDate,a.endDate)



        if salesList != None :
            salesList.append(a)
        else :
            salesList = a



        del a,tempDf
        gc.collect()

    #salesList.sort(key =lambda a: (a.server,a.category))
    return salesList


def write_data_cashshop(salesList : list[Sales]):
    totalResult = pd.DataFrame()
#print(len(salesList))

    salesList.sort(key =lambda a: (a.server,a.category))
    curRow = 0
    count = 0
    tqdmCount0=0
    print("데이터 쓰는 중...")
    for y in tqdm(salesList):
        tqdmCount0+=1
        y : Sales
        count += 1
        result = pd.DataFrame()

        i = curRow
        result.loc[i,"Category1"] = y.server
        result.loc[i,"Category2"] = y.pkgName + "\n" + str(y.pkgID)
        result.loc[i,"Category3"] = "이름"
        result.loc[i,"Check List"] = y.pkgName

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "카테고리"
        result.loc[i,"Check List"] = y.category
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "상세정보"

        if len(y.itemList0) != 0 :
            desc0 = "\n".join(y.itemList0)
            desc0 = desc0.replace("다이아몬드[귀속]","다이아몬드")
            result.loc[i,"Check List"] = desc0
        else : 
            result.loc[i,"Check List"] = f'{y.pkgName} 상자[귀속]'

        i += 1
        desc1 = "\n".join(map(str, y.itemList1))
        desc1 = desc1.replace("nan\n","")
        desc1 = desc1.replace("\n","\n- ")
        result.loc[i,"Check List"] = "사용 시 다음 아이템 획득\n\n- "+desc1

        i += 1
        desc0 = desc0.replace("\n","\n- ")
        result.loc[i,"Check List"] = "<"+y.pkgName+"> 구성품 상세 정보\n- " + desc0

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        i += 1
        result.loc[i,"Category3"] = "상품슬롯"

        for item1 in y.itemList1 :
            if str(item1) != "nan" :
                result.loc[i,"Check List"] = str(item1) + " 관련 이미지 노출"
                i+=1

        if int(y.bonus) == 0 :
            result.loc[i,"Check List"] =  "마일리지 미노출"
        else :            
            result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"
        i+=1
        result.loc[i,"Check List"] = "구매 제한 : " + str(y.limit)
        i+=1
        result.loc[i,"Check List"] = "구매 가격 : " + y.price
        
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        i += 1
        result.loc[i,"Category3"] = "아이템 구매"

        if "원" in y.price or "TWD" in y.price:
            result.loc[i,"Check List"] = f"결제 모듈 내 {y.pkgName} 노출"
            i += 1
            result.loc[i,"Check List"] = f"결제 모듈 내 {y.price} 노출"
            i += 1
            result.loc[i,"Check List"] = f"결제 완료 시 보관함으로 획득"
        else :
            result.loc[i,"Check List"] = y.price + " 차감"
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        i += 1
        result.loc[i,"Category3"] = "마일리지"

        if int(y.bonus) == 0 :
            result.loc[i,"Check List"] =  "미노출"
        else :            
            result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "아이템 획득"
        result.loc[i,"Check List"] = y.pkgName + " 상자[귀속] 인벤토리 획득"
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "아이템 사용"

        for item1 in y.itemList1 :
            if str(item1) != "nan" :
                result.loc[i,"Check List"] = "- " + str(item1) + " 획득 및 사용"
                i+=1
        
        i-=1
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "구매 제한"
        result.loc[i,"Check List"] = str(y.limit) + " 구매 시 상품 슬롯 비활성화"
        i += 1
        result.loc[i,"Check List"] = "상품 슬롯 하단에 [구매 완료] 라벨 노출"
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        curRow = i
        
        #print(f'{y.pkgName} {y.salesCheck}')
        if y.salesCheck == "판매 전" or y.salesCheck == "판매 시작":
            #print("추가")
            
            totalResult = pd.concat([totalResult,result], ignore_index=True)
        #print(len(totalResult))

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                ) 


def write_data_cashshop_inspection(salesList : list[Sales]):

    salesList.sort(key =lambda a: (a.salesCheck,a.category))


    totalResult = pd.DataFrame()
    curRow = 0
    count = 0
    #tqdmCount0=0
    print("데이터 쓰는 중...")
    for y in tqdm(salesList):

        #tqdmCount0+=1
        y : Sales
        count += 1
        result = pd.DataFrame()


        if y.salesCheck == "판매 제외" or y.salesCheck == "판매 전"  :
            continue

        

        i = curRow
        result.loc[i,"Category1"] = y.server
        result.loc[i,"Category2"] = y.salesCheck
        result.loc[i,"Category3"] = y.pkgName

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    
        info_0 = f'{y.category} / {y.price} / {y.bonus} / {y.limit}'
        
        info_expired = y.endDate.strftime('%m/%d/%Y(목) 11:00 까지')

        info_1 = "\n".join(y.itemList0)
        info_1 = info_1.replace("다이아몬드[귀속]","다이아몬드")

        info_2 = "\n".join(map(str, y.itemList1))
        info_2 = info_2.replace("nan\n","")
        info_2 = info_2.replace("\n","\n- ")
        info_2 = "사용 시 다음 아이템 획득\n- "+info_2

        info_3 = f'* 상세정보 및 패키지 상자 구성품 내 [귀속] 노출 확인\n* 패키지 이미지 내 구성품 관련 이미지 노출 확인'

        result.loc[i,"Check List"] = f'{info_0}\n{info_expired}\n\n{info_1}\n\n{info_2}\n\n{info_3}'

    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
     
        result.loc[i,"ETC"] = y.pkgID


        #if y.salesCheck != "판매 제외" and y.salesCheck != "판매 전"  :
        totalResult = pd.concat([totalResult,result], ignore_index=True)

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                ) 


def postprocess_cashshop():
    wb = xl.load_workbook(xlFileName,data_only = True)
    sheetList = wb.sheetnames
    ws = wb[sheetList[0]]
    ws.column_dimensions['b'].width = 17
    ws.column_dimensions['c'].width = 17
    ws.column_dimensions['d'].width = 17
    ws.column_dimensions['e'].width = 50

    firstRow = 2
    lastRow = ws.max_row
    startRow_B =0
    startValue_B =""
    startRow_C = 0
    startRow_D = 0

    tqdmCount1 = 0
    print("엑셀 서식 처리중...")
    for i in tqdm(range(firstRow, lastRow+1)):
        tqdmCount1+=1
        #print(i)
        if (ws['b'+str(i)].value is not None) :
            if startRow_B == 0  :
                startRow_B = i
                startValue_B = ws['b'+str(i)].value
                #print(startRow_B)
            else :
                #firstTargetCell =  "C"+str(startRow_C)
                if ( ws['b'+str(i)].value != startValue_B) :
                    mergeTargetCell = "B"+str(startRow_B)+":B"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_B = ws['b'+str(i)].value
                    startRow_B = i

        if ws['c'+str(i)].value is not None:
            if startRow_C == 0 :
                startRow_C = i
                #print(startRow)
                startValue_C = ws['c'+str(i)].value
            else :
                # firstTargetCell =  "C"+str(startRow_C)
                # mergeTargetCell = "C"+str(startRow_C)+":C"+str(i-1)
                # ws.merge_cells(mergeTargetCell)
                # startRow_C = i
                if ( ws['c'+str(i)].value != startValue_C) :
                    mergeTargetCell = "c"+str(startRow_C)+":c"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_C = ws['c'+str(i)].value
                    startRow_C = i


        if ws['d'+str(i)].value is not None:
            if startRow_D == 0 :
                startRow_D = i
                #print(startRow)
            else :
                firstTargetCell =  "D"+str(startRow_D)
                mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
                ws.merge_cells(mergeTargetCell)
                startRow_D = i


        ws['b'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['b'+str(i)].font = Font(size = 9, bold = True)
        ws['c'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['c'+str(i)].font = Font(size = 9, bold = True)
        ws['d'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['d'+str(i)].font = Font(size = 9, bold = True)
        ws['e'+str(i)].alignment = Alignment(
            horizontal='left'
            ,vertical='top'
            ,wrap_text=True)
        ws['e'+str(i)].font = Font(size = 9, bold = False)
        
        #ws['e'+str(i)].value = process_temp_str(str(ws['e'+str(i)].value))


    #예외 마지막 셀병합
    ws.merge_cells("B"+str(startRow_B)+":B"+str(lastRow))
    ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
    ws.merge_cells("D"+str(startRow_D)+":D"+str(lastRow))

    #ws = highlight_belonging(ws)
    #ws = find_and_replace(ws,"귀속","귀속")
    ws = highlight_star_cells(ws)
    wb.save(xlFileName)


def process_temp_str(temp_str):
    # Define the font style for the red asterisk
    red_asterisk_font = Font(color=Color('FF0000'), bold=True)
    # Define the font style for the rest of the cell
    normal_font = Font(color=Color('000000'))

    # Split the temp_str value by newline characters
    temp_str_lines = temp_str.split("\n")

    # Process each line of the temp_str value
    result_lines = []
    for line in temp_str_lines:
        # Check if the line starts with an asterisk
        if line.startswith("*"):
            # Apply the red font to the asterisk and the rest of the line
            result_lines.append((line[:1] + line[1:].strip(), red_asterisk_font))
        else:
            # Apply the normal font to the entire line
            result_lines.append((line, normal_font))

    # Join the processed lines and return the result
    #return "\n".join(result_lines)
    #print(result_lines)
    return "\n".join([line[0] for line in result_lines])


def highlight_belonging(ws):
    # Define the font style for the highlighted text
    highlight_font = Font(color=Color('FF0000'), bold=True)

    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell contains the word "belonging"
            if "귀속" in str(cell.value).lower():
                # Split the cell value by the word "belonging"
                parts = str(cell.value).lower().split("귀속")
                # Concatenate the parts with the highlighted "belonging" in between
                highlighted_value = "귀속".join([part.strip() for part in parts])
                # Apply the highlight font to the "belonging" text
                cell.font = highlight_font
                # Set the cell value to the highlighted value
                cell.value = highlighted_value

    # Auto-fit the columns to adjust to the new cell values
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].auto_size = True

    # Return the worksheet object
    return ws


def find_and_replace(ws, target_str, replace_str, font=None):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and target_str in str(cell.value):
                old_value = str(cell.value)
                new_value = old_value.replace(target_str, replace_str)
                cell.value = new_value
                
                if font:
                    font_obj = Font(name=font, bold=True, color="FF0000")
                    start_index = new_value.find(replace_str)
                    end_index = start_index + len(replace_str)
                    cell.font = Font(color="000000")
                    cell.font = font_obj
                    if start_index > 0:
                        cell.font = Font(color="000000", bold=False)
                    if end_index < len(new_value):
                        cell.font = Font(color="000000", bold=False)
    return ws


def highlight_star_cells(sheet):
    red_bold_font = Font(color="FF0000", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            try :
                if cell.value is not None and "*" in cell.value:
                    parts = cell.value.split("*")
                    new_parts = [f"{Font(color='black')}*{red_bold_font}".join(part.strip() for part in parts)]
                    cell.value = "\n".join(new_parts)
                    cell.font = red_bold_font
            except :
                continue
    return sheet


if __name__ == "__main__":


    #print("┃  R2M CASH SHOP CL MAKER  ┃")
    print("체크리스트 타입 입력 :")
    print("0:TC, 1:점검")
    fileType = input(">:")
    print("데이터파일명 입력 :")
    print("0:국내, 1:대만")
    fileName = input(">:")
    #fileName = ""
    if fileName == "0":
        fileName = "유료상점DATA_KR.xlsx"
    elif fileName == "1":
        fileName = "유료상점DATA_TW.xlsx"
    while not os.path.isfile(fileName) :
        fileName = input(">:")
        

    clType = ""
    if fileType == "0":
        clType = "TC"
        tcStartDate = input("(업데이트날짜 : YYYY-MM-DD)>: ")
        if tcStartDate == "" :
            tcStartDate = "2000-01-01"
        
    elif fileType == "1":
        clType = "정기점검"



#region basic Info

    cashShopDir = f"./CL_CashShop_{clType}"
    if not os.path.isdir(cashShopDir) :
        os.mkdir(cashShopDir)
    tempDir = "./temp"
    if not os.path.isdir(tempDir) :
        os.mkdir(tempDir)

    xlFileName = f"./CL_CashShop_{clType}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"
    tempCsvName = f"./temp/tempCsv.csv"

#endregion











    if fileType == "0":
        salesList = extract_data_cashshop(fileName)
        write_data_cashshop(salesList)
        postprocess_cashshop()
    elif fileType == "1":
        salesList = extract_data_cashshop(fileName)
        write_data_cashshop_inspection(salesList)
        postprocess_cashshop()


    print("생성완료")
    input("종료하려면 엔터키 입력...")